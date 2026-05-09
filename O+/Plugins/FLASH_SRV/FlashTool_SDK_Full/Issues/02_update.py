#!/usr/bin/python3

import json
import os

import openpyxl

from pathDef import SDK_ISSUES_SUPPORT_LANG_LIST, getIssuesFileByLang, wiriteDataToFile, IsEmptyString


def excel_to_json(excel_file:str=os.path.join(os.getcwd(), "allIssues.xlsx")):
    book = openpyxl.load_workbook(excel_file)
    work_sheet = book["故障树错误码"]

    row_count = work_sheet.max_row
    col_count = work_sheet.max_column

    cur_row = 2
    cur_col = 1

    json_data = []
    while cur_row <= row_count:
        issues_item = {}
        issues_item['code'] = work_sheet.cell(row=cur_row, column=cur_col).value
        json_data.append(issues_item)
        cur_row += 1

    cur_col = 2
    while cur_col <= col_count:
        lang, file_path = getIssuesFileByLang(SDK_ISSUES_SUPPORT_LANG_LIST[cur_col-2])
        print("lang={}, file_path={}".format(lang, file_path))

        cur_row = 2
        while cur_row <= row_count:
            json_data[cur_row-2]['text'] = ''
            text = work_sheet.cell(row=cur_row, column=cur_col).value
            if IsEmptyString(text) == False:
                text = text.replace("\\n", "\n")
                text = text.replace("\n\n", "\n")
            else:
                text = ''
            json_data[cur_row-2]['text'] = text
            cur_row +=1

        wiriteDataToFile(file_path, json_data)
        cur_col += 1


if __name__ == "__main__":
    excel_to_json()
    pass